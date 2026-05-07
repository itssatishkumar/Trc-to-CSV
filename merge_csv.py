import os
import math
from typing import Iterable, List
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import csv

OUTPUT_FILE = "merged.csv"

def _detect_and_strip_unit_row(df: pd.DataFrame):
    """Detect unit row and remove it."""

    if df.empty:
        return {}, df

    first_row = df.iloc[0]
    is_unit_row = False

    time_col = None
    if "Time (s)" in df.columns:
        time_col = "Time (s)"
    elif "Time" in df.columns:
        time_col = "Time"

    if time_col is not None:
        try:
            float(str(first_row[time_col]))
        except (TypeError, ValueError):
            is_unit_row = True

    units = {}

    if is_unit_row:
        for col in df.columns:
            val = first_row[col]
            if pd.notna(val) and str(val).strip() != "":
                units[col] = str(val)

        df = df.iloc[1:].reset_index(drop=True)

    return units, df


def merge_csv_files(
    csv_files: Iterable[str],
    output_file: str = OUTPUT_FILE,
    open_after: bool = False,
    row_limit: int | None = None,
):

    csv_list: List[str] = list(csv_files)

    if not csv_list:
        raise RuntimeError("No CSV files provided for merge")

    dataframes: List[pd.DataFrame] = []
    all_units = {}

    for path in csv_list:

        if not os.path.exists(path):
            print(f"Warning: CSV file not found, skipping: {path}")
            continue

        df = pd.read_csv(path, dtype=str)

        if df.empty:
            continue

        units, df_no_units = _detect_and_strip_unit_row(df)

        # -----------------------------
        # REMOVE FIRST 25 ROWS FROM EACH CSV
        # -----------------------------
        if len(df_no_units) > 25:
            df_no_units = df_no_units.iloc[25:].reset_index(drop=True)
        else:
            df_no_units = df_no_units.iloc[0:0]

        all_units.update({k: v for k, v in units.items() if v is not None})
        dataframes.append(df_no_units)

    if not dataframes:
        raise RuntimeError("No non-empty CSV data to merge")

    # -----------------------------
    # Build column union
    # -----------------------------

    all_columns = []
    seen = set()

    for df in dataframes:
        for col in df.columns:
            if col not in seen:
                seen.add(col)
                all_columns.append(col)

    # Remove Time(s) if DATE + TIME exist
    if "DATE" in seen and "TIME" in seen:
        all_columns = [c for c in all_columns if c != "Time (s)"]

    elif "Time (s)" in seen:
        all_columns = ["Time (s)"] + [c for c in all_columns if c != "Time (s)"]

    elif "Time" in seen:
        all_columns = ["Time"] + [c for c in all_columns if c != "Time"]

    normalized = [df.reindex(columns=all_columns) for df in dataframes]

    merged_df = pd.concat(normalized, ignore_index=True)

    # -----------------------------
    # Add unit row
    # -----------------------------

    if all_units:
        unit_row = [all_units.get(col, "") for col in all_columns]
        units_df = pd.DataFrame([unit_row], columns=all_columns)
        final_df = pd.concat([units_df, merged_df], ignore_index=True)

    else:
        final_df = merged_df

    # -----------------------------
    # Remove completely empty rows
    # -----------------------------

    tmp = final_df.replace(r"^\s*$", pd.NA, regex=True)
    final_df = tmp.dropna(how="all").reset_index(drop=True)

    # -----------------------------
    # Remove columns with no data
    # -----------------------------

    if len(final_df) > 1:
        data_only = final_df.iloc[1:]
        tmp_cols = data_only.replace(r"^\s*$", pd.NA, regex=True)
        cols_to_keep = ~tmp_cols.isna().all(axis=0)

        if "DATE" in final_df.columns:
            cols_to_keep["DATE"] = True

        if "TIME" in final_df.columns:
            cols_to_keep["TIME"] = True

        if "Time (s)" in final_df.columns:
            cols_to_keep["Time (s)"] = True

        final_df = final_df.loc[:, cols_to_keep]

    # -----------------------------
    # Split large files
    # -----------------------------

    if row_limit is not None and row_limit > 0:

        total_rows = len(final_df)
        total_parts = math.ceil(total_rows / row_limit)
        base, ext = os.path.splitext(output_file)
        output_paths: List[str] = []

        print(
            f"Writing merged data to CSV in {total_parts} part(s) "
            f"(max {row_limit} rows per file)..."
        )

        for i in range(total_parts):

            start = i * row_limit
            end = (i + 1) * row_limit
            chunk = final_df.iloc[start:end]
            suffix = "" if i == 0 else f"_part{i+1}"
            path = f"{base}{suffix}{ext}"
            chunk.to_csv(path, index=False)
            output_paths.append(path)
            print(f"✔ Saved: {path} ({len(chunk)} rows)")

        if open_after and output_paths:
            try:
                os.startfile(output_paths[0])
            except Exception as e:
                print(f"Could not open file: {e}")

        print(f"Merged {len(csv_list)} CSV files into {len(output_paths)} output file(s).")
        xlsx_paths = [csv_to_xlsx(p) for p in output_paths]
        return xlsx_paths

    # -----------------------------
    # Write single CSV
    # -----------------------------

    final_df.to_csv(output_file, index=False)
    xlsx_path = csv_to_xlsx(output_file)
    print(f"Merged {len(csv_list)} CSV files into {output_file}")

    if open_after:
        try:
            os.startfile(output_file)
        except Exception as e:
            print(f"Could not open file: {e}")
    return [xlsx_path]

def csv_to_xlsx(csv_path: str):
    xlsx_path = os.path.splitext(csv_path)[0] + ".xlsx"

    wb = Workbook()
    ws = wb.active

    with open(csv_path, newline='', encoding='utf-8') as fh:
        reader = list(csv.reader(fh))

    headers = reader[0]

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="00B050", end_color="00B050", fill_type="solid")
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)

    border = Border(
        left=Side(style='thin', color="000000"),
        right=Side(style='thin', color="000000"),
        top=Side(style='thin', color="000000"),
        bottom=Side(style='thin', color="000000"),
    )

    red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
    white_bold = Font(color="FFFFFF", bold=True)

    bms_idx = headers.index("BMS_State") if "BMS_State" in headers else None
    veh_idx = headers.index("VehicleState") if "VehicleState" in headers else None

    for r_idx, row in enumerate(reader, start=1):
        for c_idx, val in enumerate(row, start=1):
            val = val.strip()

            try:
                num = float(val)
                val = round(num, 2)
            except:
                pass

            cell = ws.cell(row=r_idx, column=c_idx, value=val)

            if r_idx == 1:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = center
            else:
                cell.alignment = center

                if val == "1":
                    cell.fill = red_fill
                    cell.font = white_bold

                if bms_idx is not None and c_idx - 1 == bms_idx:
                    cell.font = white_bold

                    if val == "Active":
                        cell.fill = PatternFill(start_color="00B050", end_color="00B050", fill_type="solid")
                    elif val == "Error":
                        cell.fill = red_fill
                    elif val == "Ready":
                        cell.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
                        cell.font = Font(bold=True, color="000000")
                    elif val == "Precharge":
                        cell.fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
                        cell.font = Font(bold=True, color="000000")
                    elif val == "INIT":
                        cell.fill = PatternFill(start_color="00B0F0", end_color="00B0F0", fill_type="solid")

                if veh_idx is not None and c_idx - 1 == veh_idx:
                    if val == "Drive":
                        cell.fill = PatternFill(start_color="00B0F0", end_color="00B0F0", fill_type="solid")
                        cell.font = white_bold
                    elif val == "Off":
                        cell.fill = red_fill
                        cell.font = white_bold

            cell.border = border

    for r in range(1, ws.max_row + 1):
        ws.row_dimensions[r].height = 20 if r == 1 else 14.4

    ws.freeze_panes = "A2"

    for col in ws.columns:
        max_len = 0
        col_letter = col[0].column_letter
        for cell in col:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max_len + 2, 50)

    wb.save(xlsx_path)
    return xlsx_path

if __name__ == "__main__":

    import tkinter as tk
    from tkinter import filedialog

    root = tk.Tk()
    root.withdraw()
    root.update()

    csv_files = filedialog.askopenfilenames(
        title="Select CSV files to merge",
        filetypes=[("CSV files", "*.csv")]
    )

    root.destroy()
    csv_files = list(csv_files)
    if not csv_files:
        raise RuntimeError("No CSV files selected")
    merge_csv_files(csv_files, OUTPUT_FILE, open_after=True)
